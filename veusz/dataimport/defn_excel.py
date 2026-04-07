#    Copyright (C) 2026 M. Ignacio Monge Garcia
#
#    This file is part of Plotex (based on Veusz).
#
#    Plotex is free software: you can redistribute it and/or modify it
#    under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 2 of the License, or
#    (at your option) any later version.
#
##############################################################################

"""Excel (.xlsx) data import definitions."""

from .. import qtall as qt
from .. import document
from .. import utils
from . import base

def _(text, disambiguation=None, context='Import'):
    return qt.QCoreApplication.translate(context, text, disambiguation)

class ImportParamsExcel(base.ImportParamsBase):
    """Parameters for Excel import."""

    defaults = {
        'sheet': '',
        'headerrow': True,
        'skiprows': 0,
    }
    defaults.update(base.ImportParamsBase.defaults)


class OperationDataImportExcel(base.OperationDataImportBase):
    """Operation to import data from an Excel file."""

    descr = _('import Excel data')

    def _readXlsx(self, p):
        """Read .xlsx/.xlsm with openpyxl."""
        try:
            import openpyxl
        except ImportError:
            raise base.ImportingError(
                _('openpyxl module is required for .xlsx import'))
        wb = openpyxl.load_workbook(p.filename, data_only=True, read_only=True)
        if p.sheet and p.sheet in wb.sheetnames:
            ws = wb[p.sheet]
        else:
            ws = wb.active
        allrows = []
        for row in ws.iter_rows(values_only=True):
            allrows.append(row)
        wb.close()
        return allrows

    def _readXls(self, p):
        """Read .xls with xlrd."""
        try:
            import xlrd
        except ImportError:
            raise base.ImportingError(
                _('xlrd module is required for .xls import'))
        wb = xlrd.open_workbook(p.filename)
        if p.sheet and p.sheet in wb.sheet_names():
            ws = wb.sheet_by_name(p.sheet)
        else:
            ws = wb.sheet_by_index(0)
        allrows = []
        for i in range(ws.nrows):
            allrows.append(tuple(ws.cell_value(i, c) for c in range(ws.ncols)))
        return allrows

    def doImport(self):
        """Perform the import."""
        p = self.params
        is_xls = p.filename.lower().endswith('.xls')

        try:
            if is_xls:
                try:
                    allrows = self._readXls(p)
                except Exception:
                    # fallback: might be xlsx renamed as .xls
                    allrows = self._readXlsx(p)
            else:
                allrows = self._readXlsx(p)
        except base.ImportingError:
            raise
        except Exception as e:
            raise base.ImportingError(
                _('Error opening Excel file: %s') % str(e))

        if not allrows:
            return

        # skip rows
        if p.skiprows > 0:
            allrows = allrows[p.skiprows:]
        if not allrows:
            return

        LF = LinkedFileExcel(p) if p.linked else None
        self.outdatasets = base.rows_to_datasets(
            allrows, p.headerrow, p.prefix, p.suffix, LF)


class LinkedFileExcel(base.LinkedFileBase):
    """Represents a linked Excel file for re-reading."""

    def createOperation(self):
        return OperationDataImportExcel

    def saveToFile(self, fileobj, relpath=None):
        self._saveHelper(
            fileobj,
            'ImportFileExcel',
            ('filename',),
            renameparams={'prefix': 'dsprefix', 'suffix': 'dssuffix'},
            relpath=relpath)


def ImportFileExcel(comm, filename, sheet='', headerrow=True,
                    skiprows=0, dsprefix='', dssuffix='',
                    renames=None, linked=False, encoding='utf_8'):
    """Import data from an Excel (.xlsx) file.

    sheet: name of sheet to import (default: active sheet)
    headerrow: if True, first row contains column names
    skiprows: number of rows to skip at top
    dsprefix: prefix for dataset names
    dssuffix: suffix for dataset names
    linked: if True, link to file for automatic updates
    """

    realfilename = comm.findFileOnImportPath(filename)
    params = ImportParamsExcel(
        filename=realfilename,
        sheet=sheet,
        headerrow=headerrow,
        skiprows=skiprows,
        prefix=dsprefix,
        suffix=dssuffix,
        renames=renames,
        linked=linked,
        encoding=encoding,
    )
    op = OperationDataImportExcel(params)
    comm.document.applyOperation(op)
    if comm.verbose:
        print("Imported datasets: %s" % ', '.join(op.outnames))
    return op.outnames

document.registerImportCommand('ImportFileExcel', ImportFileExcel)
