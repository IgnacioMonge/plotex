#    Copyright (C) 2026 M. Ignacio Monge Garcia
#
#    This file is part of Plotex (based on Veusz).
#
#    Plotex is free software: you can redistribute it and/or modify it
#    under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 2 of the License, or
#    (at your option) any later version.
#
##############################################################################

"""Excel (.xlsx) import dialog tab."""

from .. import qtall as qt
from .. import utils
from ..dialogs import importdialog
from . import defn_excel

def _(text, disambiguation=None, context='Import'):
    return qt.QCoreApplication.translate(context, text, disambiguation)


# ── Thread-safe standalone readers (no UI access) ───────────────

def _read_xlsx_preview(filename, sheet_name='', skiprows=0, max_rows=20):
    """Read .xlsx/.xlsm with openpyxl. Thread-safe."""
    import openpyxl
    wb = openpyxl.load_workbook(filename, data_only=True, read_only=True)
    sheets = wb.sheetnames
    ws = wb[sheet_name] if sheet_name and sheet_name in sheets else wb.active
    rows = []
    maxcols = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skiprows:
            continue
        if len(rows) >= max_rows:
            break
        rowdata = list(row)
        maxcols = max(maxcols, len(rowdata))
        rows.append(rowdata)
    wb.close()
    return sheets, rows, maxcols


def _read_xls_preview(filename, sheet_name='', skiprows=0, max_rows=20):
    """Read .xls with xlrd. Thread-safe."""
    import xlrd
    wb = xlrd.open_workbook(filename)
    sheets = wb.sheet_names()
    ws = (wb.sheet_by_name(sheet_name)
          if sheet_name and sheet_name in sheets
          else wb.sheet_by_index(0))
    rows = []
    maxcols = 0
    for i in range(ws.nrows):
        if i < skiprows:
            continue
        if len(rows) >= max_rows:
            break
        rowdata = [ws.cell_value(i, c) for c in range(ws.ncols)]
        maxcols = max(maxcols, len(rowdata))
        rows.append(rowdata)
    return sheets, rows, maxcols


def _read_excel_preview(filename, sheet_name='', skiprows=0, max_rows=20):
    """Read Excel file for preview. Tries xls then xlsx for .xls files."""
    is_xls = filename.lower().endswith('.xls')
    if is_xls:
        try:
            return _read_xls_preview(filename, sheet_name, skiprows, max_rows)
        except Exception:
            return _read_xlsx_preview(filename, sheet_name, skiprows, max_rows)
    return _read_xlsx_preview(filename, sheet_name, skiprows, max_rows)


# ── Dialog tab ──────────────────────────────────────────────────

class ImportTabExcel(importdialog.ImportTab):
    """Tab for importing Excel files."""

    resource = 'import_excel.ui'
    filetypes = ('.xlsx', '.xlsm', '.xls')
    filefilter = _('Excel files')
    handles_own_import_wrapper = True

    def loadUi(self):
        importdialog.ImportTab.loadUi(self)
        self.excelsheetcombo.currentIndexChanged.connect(
            self.dialog.slotUpdatePreview)
        self.excelheadercheck.stateChanged.connect(
            self.dialog.slotUpdatePreview)
        self.excelskiprowsspin.valueChanged.connect(
            self.dialog.slotUpdatePreview)

        self.excelheadercheck.default = True
        self.excelskiprowsspin.default = 0
        self._preview_thread = None

    def reset(self):
        self.excelsheetcombo.clear()
        self.excelheadercheck.setChecked(True)
        self.excelskiprowsspin.setValue(0)

    def doPreview(self, filename, encoding):
        """Start async preview of Excel file."""

        t = self.previewtableexcel
        t.clear()
        t.setColumnCount(1)
        t.setRowCount(1)
        t.setItem(0, 0, qt.QTableWidgetItem(_("Loading preview...")))

        if not filename or not any(
                filename.lower().endswith(ext) for ext in self.filetypes):
            t.clear()
            t.setColumnCount(0)
            t.setRowCount(0)
            return False

        # capture UI state before starting thread
        sheet_name = (self.excelsheetcombo.currentText()
                      if self.uiloaded else '')
        skiprows = (self.excelskiprowsspin.value()
                    if self.uiloaded else 0)

        # cancel previous thread
        if self._preview_thread is not None and self._preview_thread.isRunning():
            self._preview_thread.wait(500)

        self._preview_thread = importdialog.PreviewThread(
            _read_excel_preview, filename, sheet_name, skiprows)
        self._preview_thread.sigResult.connect(self._onPreviewReady)
        self._preview_thread.start()
        return False  # not ready yet; _onPreviewReady will enable import

    def _onPreviewReady(self, result):
        """Called on main thread when preview data is ready."""
        t = self.previewtableexcel
        t.clear()

        if result is None:
            t.setColumnCount(0)
            t.setRowCount(0)
            return

        sheets, rows, maxcols = result

        # update sheet selector
        current = self.excelsheetcombo.currentText()
        self.excelsheetcombo.blockSignals(True)
        self.excelsheetcombo.clear()
        self.excelsheetcombo.addItems(sheets)
        if current in sheets:
            self.excelsheetcombo.setCurrentText(current)
        self.excelsheetcombo.blockSignals(False)

        if not rows:
            t.setColumnCount(0)
            t.setRowCount(0)
            return

        # fill preview table
        t.setColumnCount(maxcols)
        t.setRowCount(len(rows))

        hasheader = self.excelheadercheck.isChecked()
        for r, rowdata in enumerate(rows):
            for c in range(maxcols):
                val = rowdata[c] if c < len(rowdata) else None
                text = str(val) if val is not None else ''
                item = qt.QTableWidgetItem(text)
                if r == 0 and hasheader:
                    f = item.font()
                    f.setBold(True)
                    item.setFont(f)
                t.setItem(r, c, item)

        if hasheader and rows:
            headers = []
            for c in range(maxcols):
                val = rows[0][c] if c < len(rows[0]) else None
                headers.append(str(val) if val is not None else '')
            t.setHorizontalHeaderLabels(headers)

        self.dialog.filepreviewokay = True
        self.dialog.enableDisableImport()

    def doImport(self, doc, filename, linked, encoding, prefix, suffix,
                 tags):
        """Import Excel data with threaded file I/O."""

        sheet = self.excelsheetcombo.currentText()
        headerrow = self.excelheadercheck.isChecked()
        skiprows = self.excelskiprowsspin.value()

        params = defn_excel.ImportParamsExcel(
            filename=filename,
            sheet=sheet,
            headerrow=headerrow,
            skiprows=skiprows,
            prefix=prefix,
            suffix=suffix,
            tags=tags,
            linked=linked,
            encoding=encoding,
        )
        op = defn_excel.OperationDataImportExcel(params)

        # Phase 1: heavy I/O in background thread with progress dialog
        progress = qt.QProgressDialog(
            _("Reading Excel file..."), None, 0, 0, self.dialog)
        progress.setWindowModality(qt.Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        qt.QApplication.processEvents()

        worker = importdialog.PreviewThread(op.preloadImport)
        loop = qt.QEventLoop()
        worker.finished.connect(loop.quit)
        worker.start()
        loop.exec()
        progress.close()

        # Phase 2: apply pre-loaded datasets to document (fast, main thread)
        with doc.suspend():
            doc.applyOperation(op)

        # show results
        t = self.previewtableexcel
        lines = self.dialog.retnDatasetInfo(op.outnames, linked, filename)
        t.clear()
        t.setColumnCount(1)
        t.setRowCount(len(lines))
        for i, l in enumerate(lines):
            t.setItem(i, 0, qt.QTableWidgetItem(l))

    def isFiletypeSupported(self, ftype):
        """Is file type supported?"""
        return ftype.lower() in self.filetypes

importdialog.registerImportTab(_('E&xcel'), ImportTabExcel)
